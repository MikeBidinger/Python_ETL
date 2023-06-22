# ------------------------------------------------------------------------------
# Developer: Mike Bidinger
# Date:      2023-06-22
# Script:    Loading of data
# ------------------------------------------------------------------------------

import openpyxl
from .json_functions import read_json, write_json
from dicttoxml import dicttoxml
from xml.dom.minidom import parseString


def write_result_list(data, file_path, append=False):
    if append:
        mode = "a"
    else:
        mode = "w"
    with open(file_path, mode) as f:
        for row in data:
            f.write(row + "\n")
    return file_path


def write_xml(data, file_path, attr_type=False, append=False):
    if append:
        mode = "a"
    else:
        mode = "w"
    my_item_func = lambda x: "row"
    xml = dicttoxml(data, attr_type=attr_type, item_func=my_item_func)
    dom = parseString(xml)
    with open(file_path, mode) as f:
        f.write(dom.toprettyxml())
    return file_path


def set_columns(headers, header_row, trim=False):
    columns = []
    if headers != []:
        for col, header in enumerate(header_row):
            if trim:
                if header.strip() in headers:
                    columns.append(col)
            elif header in headers:
                columns.append(col)
    return columns


class L_Delimited:
    def __init__(self, config_data):
        self.file_path = config_data["file_path"]
        self.delimiter = config_data["delimiter"]
        self.headers = config_data["headers"]
        # self.append = config_data["append"]

    def load(self, data):
        # Set columns for given headers
        columns = set_columns(self.headers, data[0])
        # Set data delimited
        for idx, row in enumerate(data):
            if columns == []:
                data[idx] = self.delimiter.join(row)
            else:
                row_data = []
                for col in columns:
                    row_data.append(row[col])
                data[idx] = self.delimiter.join(row_data)
        # Load data to delimited file
        return write_result_list(data, self.file_path)


class L_Positional:
    def __init__(self, config_data):
        self.file_path = config_data["file_path"]
        self.positions = config_data["positions"]
        self.headers = config_data["headers"]
        # self.append = config_data["append"]
        self.lengths = self._set_lengths()

    def _set_lengths(self):
        # Set lengths of positions
        lengths = []
        for i in range(len(self.positions) - 1):
            lengths.append(self.positions[i + 1] - self.positions[i])
        lengths.append(-1)
        return lengths

    def load(self, data):
        # Set columns for given headers
        columns = set_columns(self.headers, data[0], True)
        # Set data positional
        for idx, row in enumerate(data):
            row_data = ""
            if columns == []:
                for idy, val in enumerate(row):
                    row_data += val.ljust(self.lengths[idy], " ")
            else:
                for idy, col in enumerate(columns):
                    row_data += row[col].ljust(self.lengths[idy], " ")
            data[idx] = row_data
        # Load data to delimited file
        return write_result_list(data, self.file_path)


class L_Excel:
    def __init__(self, config_data):
        self.file_path = config_data["file_path"]
        self.headers = config_data["headers"]

    def load(self, data):
        # Set columns for given headers
        columns = set_columns(self.headers, data[0])
        # Create workbook
        wb = openpyxl.Workbook()
        # Load data to sheet
        sh = wb.active
        for idx, row in enumerate(data):
            if columns == []:
                for idy, val in enumerate(row):
                    sh.cell(idx + 1, idy + 1).value = val
            else:
                for idy, col in enumerate(columns):
                    sh.cell(idx + 1, idy + 1).value = row[col]
        # Save workbook
        wb.save(self.file_path)
        return self.file_path


class L_XML:
    def __init__(self, config_data):
        self.file_path = config_data["file_path"]

    def load(self, data):
        # Set data to dictionary
        xml = []
        for row in data[1:]:
            xml.append({})
            for idy, val in enumerate(row):
                xml[-1][data[0][idy]] = val
        # Load dictionary to xml file
        return write_xml(xml, self.file_path)


class L_JSON:
    def __init__(self, config_data):
        self.file_path = config_data["file_path"]

    def load(self, data):
        # Set data to dictionary
        json = []
        for row in data[1:]:
            json.append({})
            for idy, val in enumerate(row):
                json[-1][data[0][idy]] = val
        # Load dictionary to json file
        return write_json(self.file_path, json)
