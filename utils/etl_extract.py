# ------------------------------------------------------------------------------
# Developer: Mike Bidinger
# Date:      2023-06-23
# Script:    Extraction of data
# ------------------------------------------------------------------------------

import openpyxl
import xmltodict
from collections import OrderedDict
from .json_functions import read_json


def read_text_lines(file_path, nr_lines=0, encoding=None):
    data = []
    with open(file_path, "r", encoding=encoding) as f:
        if nr_lines == 0:
            for x in f:
                data.append(x[:-1])
        else:
            for i in range(0, nr_lines):
                line = f.readline()
                if line == "":
                    break
                data.append(line[:-1])
    return data


def read_xml(file_path):
    data = ""
    with open(file_path, "r") as f:
        data = xmltodict.parse(f.read())
    return data


class E_Delimited:
    def __init__(self, config_data):
        self.file_path = config_data["file_path"]
        self.headers = config_data["headers"]
        self.delimiter = config_data["delimiter"]

    def parse(self, nr_lines=0):
        data = []
        # Get data from file
        rows = read_text_lines(self.file_path, nr_lines)
        # Add headers (optional)
        if self.headers != []:
            data.append(self.headers)
        # Parse positions for each row
        for row in rows:
            data.append(row.split(self.delimiter))
        # Return parsed data
        return data


class E_Positional:
    def __init__(self, config_data):
        self.file_path = config_data["file_path"]
        self.headers = config_data["headers"]
        self.positions = config_data["positions"]

    def parse(self, nr_lines=0):
        data = []
        # Get data from file
        rows = read_text_lines(self.file_path, nr_lines)
        # Add headers (optional)
        if self.headers != []:
            data.append(self.headers)
        # Parse positions for each row
        for row in rows:
            row_data = []
            for idx, position in enumerate(self.positions):
                if idx + 1 < len(self.positions):
                    row_data.append(row[position : self.positions[idx + 1]])
                else:
                    row_data.append(row[position:])
            data.append(row_data)
        # Return parsed data
        return data


class E_Excel:
    def __init__(self, config_data):
        self.file_path = config_data["file_path"]
        self.headers = config_data["headers"]

    def parse(self, nr_lines=0):
        data = []
        # Get data from workbook
        wb = openpyxl.load_workbook(self.file_path)
        # Get data from sheet
        sh = wb.active
        # Add headers (optional)
        if self.headers != []:
            data.append(self.headers)
        # Parse data
        if nr_lines == 0 or nr_lines > sh.max_row:
            nr_lines = sh.max_row
        for row in range(0, nr_lines):
            row_data = []
            for col in sh.iter_cols(1, sh.max_column):
                if col[row].value is None:
                    row_data.append("")
                else:
                    row_data.append(str(col[row].value))
            data.append(row_data)
        # Return parsed data
        return data


class E_XML:
    def __init__(self, config_data):
        self.file_path = config_data["file_path"]
        self.xpath = config_data["xpath"]

    def parse(self, nr_lines=0):
        data = []
        # Get data from file
        xml = read_xml(self.file_path)
        # Get data according to XPath
        loop = self.xpath.split("/")[1:]
        elements = []
        for idx, element in enumerate(loop):
            if element == "":
                # Collect all given elements of descendant
                self._loop_children(xml, loop[idx + 1], elements)
            elif element in xml:
                xml = xml[element]
        if elements != []:
            # Set element data
            xml = []
            for x in elements:
                for y in x:
                    xml.append(y)
        # Collect all headers
        headers = {}
        row_headers = []
        for row in xml:
            if type(row) != str:
                for header in row:
                    if header not in headers:
                        headers[header] = len(headers)
                        row_headers.append(header)
            else:
                headers["rows"] = len(headers)
                row_headers.append("rows")
                break
        # Parse data
        data.append(row_headers)
        for row in xml:
            if type(row) != str:
                row_data = ["" for x in headers]
                for header in row:
                    if row[header] == None:
                        row_data[headers[header]] = ""
                    else:
                        row_data[headers[header]] = row[header]
            else:
                row_data = [row]
            data.append(row_data)
        # Return parsed data
        if nr_lines != 0:
            return data[:nr_lines]
        return data

    def _loop_children(self, xml, element, elements):
        for x in xml:
            if x == element:
                elements.append(xml[x])
            elif type(xml) == type(OrderedDict()):
                if type(xml[x]) is not str and xml[x] is not None and len(xml[x]) > 0:
                    self._loop_children(xml[x], element, elements)


class E_JSON:
    def __init__(self, config_data):
        self.file_path = config_data["file_path"]
        self.xpath = config_data["xpath"]

    def parse(self, nr_lines=0):
        data = []
        # Get data from file
        json = read_json(self.file_path)
        # Get data according to XPath
        loop = self.xpath.split("/")[1:]
        elements = []
        for idx, element in enumerate(loop):
            if element == "":
                # Collect all given elements of descendant
                self._loop_children(json, loop[idx + 1], elements)
            elif element in json:
                json = json[element]
        if elements != []:
            # Set element data
            json = []
            for x in elements:
                for y in x:
                    json.append(y)
        # Collect all headers
        headers = {}
        row_headers = []
        for row in json:
            if type(row) != str:
                for header in row:
                    if header not in headers:
                        headers[header] = len(headers)
                        row_headers.append(header)
            else:
                headers["rows"] = len(headers)
                row_headers.append("rows")
                break
        # Parse data
        data.append(row_headers)
        for row in json:
            if type(row) != str:
                row_data = ["" for x in headers]
                for header in row:
                    if row[header] == None:
                        row_data[headers[header]] = ""
                    else:
                        row_data[headers[header]] = row[header]
            else:
                row_data = [row]
            data.append(row_data)
        # Return parsed data
        if nr_lines != 0:
            return data[:nr_lines]
        return data

    def _loop_children(self, json, element, elements):
        for x in json:
            if x == element:
                elements.append(json[x])
            elif type(json) == type(OrderedDict()):
                if type(json[x]) is not str and json[x] is not None and len(json[x]) > 0:
                    self._loop_children(json[x], element, elements)
